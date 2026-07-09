"""Tests for src/report_export.py — workbook structure, sheet names, representative
cell values, and the manifest each export function returns. This module only formats
already-computed envelopes, so these tests never re-derive business-rule outcomes."""

from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from src.report_export import (
    ALLOCATION_RESULT_COLUMNS,
    DRAFT_MESSAGE_COLUMNS,
    PAYMENT_AGING_ROW_COLUMNS,
    PAYMENT_DATA_ISSUE_COLUMNS,
    REMAINING_INVENTORY_COLUMNS,
    SUPPLIER_FOLLOW_UP_COLUMNS,
    VALID_ORDER_COLUMNS,
    VALIDATION_ERROR_COLUMNS,
    export_inventory_allocation_report,
    export_order_validation_report,
    export_payment_aging_report,
)
from tests.contract_fixtures import (
    ALLOCATION_RESULT_ROW_FIXTURE,
    ALLOCATION_SUMMARY_FIXTURE,
    BACKORDER_ROW_FIXTURE,
    DRAFT_MESSAGE_ROW_FIXTURE,
    PAYMENT_AGING_ROW_FIXTURE,
    PAYMENT_AGING_SUMMARY_FIXTURE,
    PAYMENT_DATA_ISSUE_ROW_FIXTURE,
    REMAINING_INVENTORY_ROW_FIXTURE,
    REPORT_MANIFEST_FIXTURES,
    SUPPLIER_FOLLOW_UP_ROW_FIXTURE,
    VALID_ORDER_ROW_FIXTURE,
    VALIDATION_ERROR_ROW_FIXTURE,
    VALIDATION_SUMMARY_FIXTURE,
)

GENERATED_AT = datetime(2026, 7, 9, 9, 15, 0, 123456)


def _sheet_names_for(report_type: str) -> list[str]:
    return next(m["sheet_names"] for m in REPORT_MANIFEST_FIXTURES if m["report_type"] == report_type)


def _load(workbook_bytes: bytes):
    return load_workbook(io.BytesIO(workbook_bytes))


def _headers(ws) -> list:
    return [cell.value for cell in ws[1]]


# --- Order validation report -------------------------------------------------


def _order_validation_result(**overrides) -> dict:
    result = {
        "summary": VALIDATION_SUMMARY_FIXTURE,
        "valid_orders": [VALID_ORDER_ROW_FIXTURE],
        "errors": [VALIDATION_ERROR_ROW_FIXTURE],
    }
    result.update(overrides)
    return result


def test_order_validation_report_sheet_names_match_manifest_fixture():
    orders_df = pd.DataFrame([{"order_id": "SO-2026-006", "sku": "OFF-CHAIR-006"}])
    workbook_bytes, manifest = export_order_validation_report(
        _order_validation_result(), orders_df, generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    expected = _sheet_names_for("order_validation")
    assert wb.sheetnames == expected
    assert manifest["sheet_names"] == expected


def test_order_validation_report_manifest_shape():
    workbook_bytes, manifest = export_order_validation_report(
        _order_validation_result(), pd.DataFrame(), generated_at=GENERATED_AT
    )
    assert set(manifest.keys()) == {"report_id", "report_type", "file_name", "generated_at", "sheet_names"}
    assert manifest["report_type"] == "order_validation"
    assert manifest["file_name"] == "order_validation_report.xlsx"
    assert manifest["report_id"] == "rpt-order_validation-20260709091500"
    assert manifest["generated_at"] == "2026-07-09T09:15:00"


def test_order_validation_report_header_rows_match_column_constants():
    workbook_bytes, _ = export_order_validation_report(
        _order_validation_result(), pd.DataFrame(), generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    assert _headers(wb["Valid Orders"]) == VALID_ORDER_COLUMNS
    assert _headers(wb["Validation Errors"]) == VALIDATION_ERROR_COLUMNS


def test_order_validation_report_representative_row_values():
    workbook_bytes, _ = export_order_validation_report(
        _order_validation_result(), pd.DataFrame(), generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    ws = wb["Valid Orders"]
    row = {header: cell for header, cell in zip(VALID_ORDER_COLUMNS, [c.value for c in ws[2]])}
    assert row["order_id"] == VALID_ORDER_ROW_FIXTURE["order_id"]
    assert row["quantity"] == VALID_ORDER_ROW_FIXTURE["quantity"]

    errors_ws = wb["Validation Errors"]
    error_row = {header: cell for header, cell in zip(VALIDATION_ERROR_COLUMNS, [c.value for c in errors_ws[2]])}
    assert error_row["error_code"] == VALIDATION_ERROR_ROW_FIXTURE["error_code"]


def test_order_validation_report_empty_errors_sheet_has_header_only():
    workbook_bytes, _ = export_order_validation_report(
        _order_validation_result(errors=[]), pd.DataFrame(), generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    ws = wb["Validation Errors"]
    assert ws.max_row == 1
    assert _headers(ws) == VALIDATION_ERROR_COLUMNS


def test_order_validation_report_original_orders_mirrors_input_dataframe():
    orders_df = pd.DataFrame(
        [
            {"order_id": "SO-2026-006", "quantity": 15},
            {"order_id": "SO-2026-007", "quantity": None},
        ]
    )
    workbook_bytes, _ = export_order_validation_report(
        _order_validation_result(), orders_df, generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    ws = wb["Original Orders"]
    assert _headers(ws) == ["order_id", "quantity"]
    assert [c.value for c in ws[2]] == ["SO-2026-006", 15]
    # openpyxl round-trips a written "" as a genuinely blank cell -> None on reload.
    assert [c.value for c in ws[3]] == ["SO-2026-007", None]


def test_order_validation_report_original_orders_empty_when_none():
    workbook_bytes, manifest = export_order_validation_report(
        _order_validation_result(), None, generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    assert "Original Orders" in wb.sheetnames
    ws = wb["Original Orders"]
    assert ws.max_row == 1
    assert manifest["sheet_names"] == _sheet_names_for("order_validation")


def test_order_validation_report_missing_notrequired_field_renders_blank():
    error_without_sku = {
        "row_number": 3,
        "error_code": "OV-001",
        "error_message": "Customer name is missing",
        "severity": "Error",
    }
    workbook_bytes, _ = export_order_validation_report(
        _order_validation_result(errors=[error_without_sku]), pd.DataFrame(), generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    ws = wb["Validation Errors"]
    row = dict(zip(VALIDATION_ERROR_COLUMNS, [c.value for c in ws[2]]))
    # openpyxl round-trips a written "" as a genuinely blank cell -> None on reload.
    assert row["sku"] is None
    assert row["order_id"] is None


def test_order_validation_report_generated_at_defaults_when_omitted():
    workbook_bytes, manifest = export_order_validation_report(_order_validation_result(), pd.DataFrame())
    assert manifest["generated_at"] != ""
    assert manifest["report_id"].startswith("rpt-order_validation-")


# --- Inventory allocation report ---------------------------------------------


def _inventory_allocation_result(**overrides) -> dict:
    result = {
        "summary": ALLOCATION_SUMMARY_FIXTURE,
        "allocation_results": [ALLOCATION_RESULT_ROW_FIXTURE],
        "backorders": [BACKORDER_ROW_FIXTURE],
        "remaining_inventory": [REMAINING_INVENTORY_ROW_FIXTURE],
        "supplier_follow_ups": [SUPPLIER_FOLLOW_UP_ROW_FIXTURE],
    }
    result.update(overrides)
    return result


def test_inventory_allocation_report_sheet_names_match_manifest_fixture():
    workbook_bytes, manifest = export_inventory_allocation_report(
        _inventory_allocation_result(), generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    expected = _sheet_names_for("inventory_allocation")
    assert wb.sheetnames == expected
    assert manifest["sheet_names"] == expected
    assert manifest["report_type"] == "inventory_allocation"
    assert manifest["file_name"] == "inventory_allocation_report.xlsx"


def test_inventory_allocation_report_header_rows_match_column_constants():
    workbook_bytes, _ = export_inventory_allocation_report(
        _inventory_allocation_result(), generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    assert _headers(wb["Allocation Results"]) == ALLOCATION_RESULT_COLUMNS
    assert _headers(wb["Backorders"]) == ALLOCATION_RESULT_COLUMNS
    assert _headers(wb["Remaining Inventory"]) == REMAINING_INVENTORY_COLUMNS
    assert _headers(wb["Supplier Follow-up"]) == SUPPLIER_FOLLOW_UP_COLUMNS


def test_inventory_allocation_report_backorders_sheet_uses_precomputed_list_as_is():
    workbook_bytes, _ = export_inventory_allocation_report(
        _inventory_allocation_result(), generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    ws = wb["Backorders"]
    row = dict(zip(ALLOCATION_RESULT_COLUMNS, [c.value for c in ws[2]]))
    assert row["order_id"] == BACKORDER_ROW_FIXTURE["order_id"]
    assert row["status"] == "Backordered"


def test_inventory_allocation_report_empty_backorders_sheet_has_header_only():
    workbook_bytes, _ = export_inventory_allocation_report(
        _inventory_allocation_result(backorders=[]), generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    ws = wb["Backorders"]
    assert ws.max_row == 1


def test_inventory_allocation_report_summary_sheet_representative_values():
    workbook_bytes, _ = export_inventory_allocation_report(
        _inventory_allocation_result(), generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    ws = wb["Allocation Summary"]
    values = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2)}
    assert values["Total Order Lines"] == ALLOCATION_SUMMARY_FIXTURE["total_order_lines"]
    assert values["Backordered Count"] == ALLOCATION_SUMMARY_FIXTURE["backordered_count"]


# --- Payment aging report -----------------------------------------------------


def _payment_aging_result(**overrides) -> dict:
    result = {
        "summary": PAYMENT_AGING_SUMMARY_FIXTURE,
        "aging_rows": [PAYMENT_AGING_ROW_FIXTURE],
        "data_issues": [PAYMENT_DATA_ISSUE_ROW_FIXTURE],
        "draft_messages": [DRAFT_MESSAGE_ROW_FIXTURE],
    }
    result.update(overrides)
    return result


def test_payment_aging_report_sheet_names_match_manifest_fixture():
    workbook_bytes, manifest = export_payment_aging_report(_payment_aging_result(), generated_at=GENERATED_AT)
    wb = _load(workbook_bytes)
    expected = _sheet_names_for("payment_aging")
    assert wb.sheetnames == expected
    assert manifest["sheet_names"] == expected
    assert manifest["report_type"] == "payment_aging"
    assert manifest["file_name"] == "payment_aging_report.xlsx"


def test_payment_aging_report_header_rows_match_column_constants():
    workbook_bytes, _ = export_payment_aging_report(_payment_aging_result(), generated_at=GENERATED_AT)
    wb = _load(workbook_bytes)
    assert _headers(wb["Follow-up List"]) == PAYMENT_AGING_ROW_COLUMNS
    assert _headers(wb["All Invoices with Aging"]) == PAYMENT_AGING_ROW_COLUMNS
    assert _headers(wb["Data Issues"]) == PAYMENT_DATA_ISSUE_COLUMNS
    assert _headers(wb["Draft Messages"]) == DRAFT_MESSAGE_COLUMNS


def test_payment_aging_report_follow_up_list_includes_watch_high_medium_low():
    aging_rows = [
        {**PAYMENT_AGING_ROW_FIXTURE, "invoice_id": "INV-WATCH", "follow_up_priority": "Watch"},
        {**PAYMENT_AGING_ROW_FIXTURE, "invoice_id": "INV-HIGH", "follow_up_priority": "High"},
        {**PAYMENT_AGING_ROW_FIXTURE, "invoice_id": "INV-MEDIUM", "follow_up_priority": "Medium"},
        {**PAYMENT_AGING_ROW_FIXTURE, "invoice_id": "INV-LOW", "follow_up_priority": "Low"},
        {**PAYMENT_AGING_ROW_FIXTURE, "invoice_id": "INV-NONE", "follow_up_priority": "None"},
    ]
    workbook_bytes, _ = export_payment_aging_report(
        _payment_aging_result(aging_rows=aging_rows), generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    ws = wb["Follow-up List"]
    invoice_ids = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert invoice_ids == ["INV-WATCH", "INV-HIGH", "INV-MEDIUM", "INV-LOW"]

    all_invoices_ws = wb["All Invoices with Aging"]
    all_invoice_ids = [row[0].value for row in all_invoices_ws.iter_rows(min_row=2)]
    assert all_invoice_ids == ["INV-WATCH", "INV-HIGH", "INV-MEDIUM", "INV-LOW", "INV-NONE"]


def test_payment_aging_report_follow_up_list_excludes_unexpected_priority_string():
    aging_rows = [{**PAYMENT_AGING_ROW_FIXTURE, "follow_up_priority": "Unexpected"}]
    workbook_bytes, _ = export_payment_aging_report(
        _payment_aging_result(aging_rows=aging_rows), generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    ws = wb["Follow-up List"]
    assert ws.max_row == 1


def test_payment_aging_report_empty_data_issues_sheet_has_header_only():
    workbook_bytes, _ = export_payment_aging_report(
        _payment_aging_result(data_issues=[]), generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    ws = wb["Data Issues"]
    assert ws.max_row == 1
    assert _headers(ws) == PAYMENT_DATA_ISSUE_COLUMNS


def test_payment_aging_report_data_issue_missing_notrequired_field_renders_blank():
    issue_without_invoice_id = {
        "error_code": "PA-006",
        "error_message": "Due date is missing",
        "severity": "Error",
    }
    workbook_bytes, _ = export_payment_aging_report(
        _payment_aging_result(data_issues=[issue_without_invoice_id]), generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    ws = wb["Data Issues"]
    row = dict(zip(PAYMENT_DATA_ISSUE_COLUMNS, [c.value for c in ws[2]]))
    # openpyxl round-trips a written "" as a genuinely blank cell -> None on reload.
    assert row["invoice_id"] is None
    assert row["customer_name"] is None


def test_payment_aging_report_summary_sheet_flattens_aging_bucket_counts():
    workbook_bytes, _ = export_payment_aging_report(_payment_aging_result(), generated_at=GENERATED_AT)
    wb = _load(workbook_bytes)
    ws = wb["Aging Summary"]
    labels = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert "Aging Bucket Counts: Current" in labels
    assert "Aging Bucket Counts: 90+ Days" in labels
    values = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2)}
    assert values["Aging Bucket Counts: 1-30 Days"] == PAYMENT_AGING_SUMMARY_FIXTURE["aging_bucket_counts"]["1-30 Days"]
    assert values["Total Invoices"] == PAYMENT_AGING_SUMMARY_FIXTURE["total_invoices"]


def test_payment_aging_report_aging_summary_follows_provided_dict_not_hardcoded_order():
    # Deliberately different keys/order/count than the old 5-key AGING_BUCKET_ORDER list
    # (Current, 1-30 Days, 31-60 Days, 61-90 Days, 90+ Days) — proves the summary writer
    # renders whatever dict it's handed, not a report_export-owned bucket list.
    custom_summary = {
        **PAYMENT_AGING_SUMMARY_FIXTURE,
        "aging_bucket_counts": {
            "91+ Days": 4,
            "Current": 1,
            "1-30 Days": 2,
            "61-90 Days": 3,
            "Zzz-Custom-Bucket": 9,
        },
    }
    workbook_bytes, _ = export_payment_aging_report(
        _payment_aging_result(summary=custom_summary), generated_at=GENERATED_AT
    )
    wb = _load(workbook_bytes)
    ws = wb["Aging Summary"]
    rows = [(row[0].value, row[1].value) for row in ws.iter_rows(min_row=2)]
    bucket_rows = [r for r in rows if r[0].startswith("Aging Bucket Counts:")]
    assert bucket_rows == [
        ("Aging Bucket Counts: 91+ Days", 4),
        ("Aging Bucket Counts: Current", 1),
        ("Aging Bucket Counts: 1-30 Days", 2),
        ("Aging Bucket Counts: 61-90 Days", 3),
        ("Aging Bucket Counts: Zzz-Custom-Bucket", 9),
    ]


def test_payment_aging_report_draft_message_cell_wraps_text():
    # DRAFT_MESSAGE_ROW_FIXTURE.message_text is a real multi-paragraph string with
    # embedded \n — must be readable when opened in Excel, not squashed onto one line.
    workbook_bytes, _ = export_payment_aging_report(_payment_aging_result(), generated_at=GENERATED_AT)
    wb = _load(workbook_bytes)
    ws = wb["Draft Messages"]
    message_text_col = DRAFT_MESSAGE_COLUMNS.index("message_text") + 1  # 1-indexed
    cell = ws.cell(row=2, column=message_text_col)
    assert cell.value == DRAFT_MESSAGE_ROW_FIXTURE["message_text"]
    assert cell.alignment.wrap_text is True
    assert cell.alignment.vertical == "top"
    column_letter = cell.column_letter
    assert ws.column_dimensions[column_letter].width > 60


def test_payment_aging_report_report_id_format():
    workbook_bytes, manifest = export_payment_aging_report(_payment_aging_result(), generated_at=GENERATED_AT)
    assert manifest["report_id"] == "rpt-payment_aging-20260709091500"
    assert manifest["generated_at"] == "2026-07-09T09:15:00"
