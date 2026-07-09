"""Excel workbook generation from already-computed business-module results.

This module performs no business calculations, no revalidation, and no allocation
or aging logic — it only formats and writes already-computed output-contract data
to .xlsx workbooks. All workbooks are built in memory (io.BytesIO) and returned as
bytes; this module never writes to disk.
"""

from __future__ import annotations

import io
import math
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.contracts import ReportManifest
from src.inventory_allocation import InventoryAllocationResult
from src.order_validation import OrderValidationResult
from src.payment_aging import PaymentAgingResult

# Excel workbook styling constants (backend .xlsx formatting only — distinct from the
# Next.js/Tailwind semantic UI token system, which these do not apply to).
HEADER_FILL_COLOR = "D9E1F2"
HEADER_FONT_BOLD = True

VALIDATION_ERROR_COLUMNS = ["row_number", "order_id", "sku", "error_code", "error_message", "severity"]
VALID_ORDER_COLUMNS = [
    "order_id",
    "order_date",
    "customer_name",
    "customer_region",
    "sku",
    "product_name",
    "quantity",
    "requested_delivery_date",
    "priority",
    "payment_terms",
    "sales_owner",
]
ALLOCATION_RESULT_COLUMNS = [
    "order_id",
    "customer_name",
    "sku",
    "product_name",
    "requested_qty",
    "allocated_qty",
    "backorder_qty",
    "warehouse",
    "status",
    "priority",
    "requested_delivery_date",
]
REMAINING_INVENTORY_COLUMNS = [
    "sku",
    "warehouse",
    "starting_available_qty",
    "allocated_qty",
    "remaining_qty",
    "reorder_point",
    "reorder_alert",
]
SUPPLIER_FOLLOW_UP_COLUMNS = ["sku", "warehouse", "remaining_qty", "reorder_point", "supplier_name", "lead_time_days"]
PAYMENT_AGING_ROW_COLUMNS = [
    "invoice_id",
    "customer_name",
    "invoice_date",
    "due_date",
    "invoice_amount",
    "paid_amount",
    "outstanding_amount",
    "days_overdue",
    "aging_bucket",
    "follow_up_priority",
    "suggested_action",
]
PAYMENT_DATA_ISSUE_COLUMNS = ["invoice_id", "customer_name", "error_code", "error_message", "severity"]
DRAFT_MESSAGE_COLUMNS = ["invoice_id", "customer_name", "outstanding_amount", "days_overdue", "message_text"]

FOLLOW_UP_PRIORITIES = {"High", "Medium", "Low", "Watch"}


def _safe_cell_value(value: Any) -> Any:
    """Normalize None/NaN/NaT/pd.NA to an empty string for a clean Excel cell."""
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return value


def _new_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _style_header_row(ws: Worksheet, num_columns: int) -> None:
    fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
    font = Font(bold=HEADER_FONT_BOLD)
    for col_index in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col_index)
        cell.font = font
        cell.fill = fill
    ws.freeze_panes = "A2"


def _autosize_columns(ws: Worksheet, headers: list[str]) -> None:
    for col_index, header in enumerate(headers, start=1):
        column_letter = ws.cell(row=1, column=col_index).column_letter
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):
            cell_value = row[0].value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def _write_detail_sheet(wb: Workbook, sheet_name: str, rows: list[dict], columns: list[str]) -> Worksheet:
    ws = wb.create_sheet(sheet_name)
    ws.append(columns)
    for row in rows:
        ws.append([_safe_cell_value(row.get(col, "")) for col in columns])
    _style_header_row(ws, len(columns))
    _autosize_columns(ws, columns)
    return ws


def _apply_wrap_text_to_column(ws: Worksheet, column_index: int, width: int = 80, row_height: float = 90) -> None:
    """Wrap long multi-line text (e.g. Draft Messages' message_text) instead of squashing
    it onto one unreadable line — used only for the one column that needs it."""
    column_letter = ws.cell(row=1, column=column_index).column_letter
    ws.column_dimensions[column_letter].width = width
    for row in ws.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
        cell = row[0]
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[cell.row].height = row_height


def _write_summary_sheet(wb: Workbook, sheet_name: str, summary: dict) -> Worksheet:
    ws = wb.create_sheet(sheet_name)
    ws.append(["Metric", "Value"])
    for key, value in summary.items():
        if isinstance(value, dict):
            for sub_key, sub_value in value.items():
                label = f"{_label(key)}: {sub_key}"
                ws.append([label, _safe_cell_value(sub_value)])
        else:
            ws.append([_label(key), _safe_cell_value(value)])
    _style_header_row(ws, 2)
    _autosize_columns(ws, ["Metric", "Value"])
    return ws


def _label(field_name: str) -> str:
    return field_name.replace("_", " ").title()


def _write_raw_dataframe_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame | None) -> Worksheet:
    ws = wb.create_sheet(sheet_name)
    if df is None or df.empty:
        columns = list(df.columns) if df is not None else []
        if columns:
            ws.append(columns)
            _style_header_row(ws, len(columns))
        return ws
    columns = list(df.columns)
    ws.append(columns)
    for _, row in df.iterrows():
        ws.append([_safe_cell_value(value) for value in row.tolist()])
    _style_header_row(ws, len(columns))
    _autosize_columns(ws, [str(c) for c in columns])
    return ws


def _build_manifest(
    report_type: str, file_name: str, sheet_names: list[str], generated_at: datetime
) -> ReportManifest:
    report_id = f"rpt-{report_type}-{generated_at:%Y%m%d%H%M%S}"
    return {
        "report_id": report_id,
        "report_type": report_type,
        "file_name": file_name,
        "generated_at": generated_at.isoformat(timespec="seconds"),
        "sheet_names": sheet_names,
    }


def _save_workbook_bytes(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_order_validation_report(
    result: OrderValidationResult,
    original_orders_df: pd.DataFrame | None = None,
    generated_at: datetime | None = None,
) -> tuple[bytes, ReportManifest]:
    """Build order_validation_report.xlsx from an already-computed OrderValidationResult."""
    effective_generated_at = generated_at or datetime.now()
    wb = _new_workbook()
    _write_summary_sheet(wb, "Summary", result["summary"])
    _write_detail_sheet(wb, "Valid Orders", result["valid_orders"], VALID_ORDER_COLUMNS)
    _write_detail_sheet(wb, "Validation Errors", result["errors"], VALIDATION_ERROR_COLUMNS)
    _write_raw_dataframe_sheet(wb, "Original Orders", original_orders_df)

    sheet_names = wb.sheetnames
    workbook_bytes = _save_workbook_bytes(wb)
    manifest = _build_manifest(
        "order_validation", "order_validation_report.xlsx", sheet_names, effective_generated_at
    )
    return workbook_bytes, manifest


def export_inventory_allocation_report(
    result: InventoryAllocationResult,
    generated_at: datetime | None = None,
) -> tuple[bytes, ReportManifest]:
    """Build inventory_allocation_report.xlsx from an already-computed InventoryAllocationResult."""
    effective_generated_at = generated_at or datetime.now()
    wb = _new_workbook()
    _write_summary_sheet(wb, "Allocation Summary", result["summary"])
    _write_detail_sheet(wb, "Allocation Results", result["allocation_results"], ALLOCATION_RESULT_COLUMNS)
    _write_detail_sheet(wb, "Backorders", result["backorders"], ALLOCATION_RESULT_COLUMNS)
    _write_detail_sheet(wb, "Remaining Inventory", result["remaining_inventory"], REMAINING_INVENTORY_COLUMNS)
    _write_detail_sheet(wb, "Supplier Follow-up", result["supplier_follow_ups"], SUPPLIER_FOLLOW_UP_COLUMNS)

    sheet_names = wb.sheetnames
    workbook_bytes = _save_workbook_bytes(wb)
    manifest = _build_manifest(
        "inventory_allocation", "inventory_allocation_report.xlsx", sheet_names, effective_generated_at
    )
    return workbook_bytes, manifest


def export_payment_aging_report(
    result: PaymentAgingResult,
    generated_at: datetime | None = None,
) -> tuple[bytes, ReportManifest]:
    """Build payment_aging_report.xlsx from an already-computed PaymentAgingResult."""
    effective_generated_at = generated_at or datetime.now()
    follow_up_rows = [
        row for row in result["aging_rows"] if row["follow_up_priority"] in FOLLOW_UP_PRIORITIES
    ]

    wb = _new_workbook()
    _write_summary_sheet(wb, "Aging Summary", result["summary"])
    _write_detail_sheet(wb, "Follow-up List", follow_up_rows, PAYMENT_AGING_ROW_COLUMNS)
    _write_detail_sheet(wb, "All Invoices with Aging", result["aging_rows"], PAYMENT_AGING_ROW_COLUMNS)
    _write_detail_sheet(wb, "Data Issues", result["data_issues"], PAYMENT_DATA_ISSUE_COLUMNS)
    draft_messages_ws = _write_detail_sheet(wb, "Draft Messages", result["draft_messages"], DRAFT_MESSAGE_COLUMNS)
    _apply_wrap_text_to_column(draft_messages_ws, DRAFT_MESSAGE_COLUMNS.index("message_text") + 1)

    sheet_names = wb.sheetnames
    workbook_bytes = _save_workbook_bytes(wb)
    manifest = _build_manifest(
        "payment_aging", "payment_aging_report.xlsx", sheet_names, effective_generated_at
    )
    return workbook_bytes, manifest
